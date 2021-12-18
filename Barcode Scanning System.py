import cv2
import numpy as np
from pyzbar.pyzbar import decode
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
wb.active.title="Barcode data"
sheet = wb.active

sheet['A1'].value="Sr.No"
sheet['A1'].alignment = Alignment(horizontal='center')
sheet['B1'].value="Data"
sheet['B1'].alignment = Alignment(horizontal='center')
sheet['C1'].value="Type"
sheet['C1'].alignment = Alignment(horizontal='center')

wb.save("C:\\Users\\M\\Desktop\\S.P.I.T\\SEM 5\\DSP\\MINI PROJECT\\Python code\\Data.xlsx")  # add location/path where u want to save excel file

bardata = []
code_type =[]

cap =cv2.VideoCapture(0)
cap.set(3,640)
cap.set(4,480)

while True:
    success, img = cap.read()
    for barcode in decode(img):
        mydata= barcode.data.decode('utf-8')
        codetype= barcode.type
        
        if mydata in bardata:
            print("done")
        else:
            bardata.append(mydata)
            code_type.append(codetype)
        print(bardata)

        length = len(bardata)
        length1= len(code_type)
      
        srno = set(range(1,length+2))
        #print(srno)
        Srno = list(srno)

        for i in range(2,length+2):
            for j in range(1,2):
                cellref=sheet.cell(i,j)
                cellref.alignment = Alignment(horizontal='center')
                cellref.value= Srno[i-2]
                
        for i in range(2,length+2):
            for j in range(2,3):
                cellref=sheet.cell(i,j)
                cellref.alignment = Alignment(horizontal='center')
                cellref.value=bardata[i-2] 
                
        for i in range(2,length+2):
            for j in range(3,4):
                cellref=sheet.cell(i,j)
                cellref.alignment = Alignment(horizontal='center')
                cellref.value=code_type[i-2] 
                wb.save("C:\\Users\\M\\Desktop\\S.P.I.T\\SEM 5\\DSP\\MINI PROJECT\\Python code\\Data.xlsx")  # add location/path where u want to save excel file      

        pts =np.array([barcode.polygon],np.int32)
        pts =pts.reshape((-1,1,2))
        cv2.polylines(img,[pts],True,(255,0,255),5)
        pts2 = barcode.rect
        cv2.putText(img,mydata,(pts2[0],pts2[1]),cv2.FONT_HERSHEY_SIMPLEX,0.9,(255,0,255),2)

    cv2.imshow('Scanner',img)
    cv2.waitKey(1)

